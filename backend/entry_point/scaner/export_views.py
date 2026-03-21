import io
import logging
from datetime import timedelta

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from drf_spectacular.utils import extend_schema, OpenApiResponse

from user.models import User
from scaner.models import WeeklyRecord, MonthlyRecord, YearlyRecord

logger = logging.getLogger('scaner')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
center = Alignment(horizontal='center', vertical='center')


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def _style_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left', vertical='center')


def _build_workbook():
    """Генерирует Excel-книгу со статистикой сотрудников."""
    wb = Workbook()
    today = timezone.now().date()

    users = User.objects.filter(is_admin=False).order_by('surname', 'name')

    # ===== Лист «Неделя» =====
    ws_week = wb.active
    ws_week.title = 'Неделя'
    week_start = today - timedelta(days=today.weekday())

    week_headers = ['№', 'Фамилия', 'Имя', 'Отчество', 'Email',
                    'Начало недели', 'Конец недели', 'Дней отработано', 'Часов отработано']
    ws_week.append(week_headers)
    _style_header(ws_week, 1, len(week_headers))

    for idx, user in enumerate(users, start=1):
        rec = WeeklyRecord.objects.filter(user=user, week_start=week_start).first()
        ws_week.append([
            idx,
            user.surname,
            user.name,
            user.patronymic or '',
            user.email,
            str(week_start),
            str(week_start + timedelta(days=6)),
            rec.days_worked if rec else 0,
            rec.total_hours if rec else 0,
        ])
        _style_row(ws_week, idx + 1, len(week_headers))

    for col_idx in range(1, len(week_headers) + 1):
        ws_week.column_dimensions[ws_week.cell(1, col_idx).column_letter].width = 18

    # ===== Лист «Месяц» =====
    ws_month = wb.create_sheet('Месяц')
    month_headers = ['№', 'Фамилия', 'Имя', 'Отчество', 'Email',
                     'Начало месяца', 'Конец месяца', 'Дней отработано', 'Часов отработано']
    ws_month.append(month_headers)
    _style_header(ws_month, 1, len(month_headers))

    for idx, user in enumerate(users, start=1):
        rec = MonthlyRecord.objects.filter(user=user, year=today.year, month=today.month).first()
        # Если месяц ещё не финализирован — суммируем из недельных записей
        if rec:
            days = rec.days_worked
            hours = rec.total_hours
            sd = str(rec.start_date)
            ed = str(rec.end_date)
        else:
            # Собираем из недельных за текущий месяц
            weeks = WeeklyRecord.objects.filter(
                user=user,
                week_start__year=today.year, week_start__month=today.month,
            )
            days = sum(w.days_worked for w in weeks)
            hours = round(sum(w.total_seconds for w in weeks) / 3600, 2)
            import calendar
            sd = str(today.replace(day=1))
            last_day = calendar.monthrange(today.year, today.month)[1]
            ed = str(today.replace(day=last_day))

        ws_month.append([idx, user.surname, user.name, user.patronymic or '',
                         user.email, sd, ed, days, hours])
        _style_row(ws_month, idx + 1, len(month_headers))

    for col_idx in range(1, len(month_headers) + 1):
        ws_month.column_dimensions[ws_month.cell(1, col_idx).column_letter].width = 18

    # ===== Лист «Год» =====
    ws_year = wb.create_sheet('Год')
    year_headers = ['№', 'Фамилия', 'Имя', 'Отчество', 'Email',
                    'Начало года', 'Конец года', 'Дней отработано', 'Часов отработано']
    ws_year.append(year_headers)
    _style_header(ws_year, 1, len(year_headers))

    for idx, user in enumerate(users, start=1):
        rec = YearlyRecord.objects.filter(user=user, year=today.year).first()
        if rec:
            days = rec.days_worked
            hours = rec.total_hours
            sd = str(rec.start_date)
            ed = str(rec.end_date)
        else:
            # Собираем из месячных + текущих недельных
            months = MonthlyRecord.objects.filter(user=user, year=today.year)
            weeks = WeeklyRecord.objects.filter(user=user, week_start__year=today.year, is_finalized=False)
            days = sum(m.days_worked for m in months) + sum(w.days_worked for w in weeks)
            secs = sum(m.total_seconds for m in months) + sum(w.total_seconds for w in weeks)
            hours = round(secs / 3600, 2)
            sd = str(today.replace(month=1, day=1))
            ed = str(today.replace(month=12, day=31))

        ws_year.append([idx, user.surname, user.name, user.patronymic or '',
                        user.email, sd, ed, days, hours])
        _style_row(ws_year, idx + 1, len(year_headers))

    for col_idx in range(1, len(year_headers) + 1):
        ws_year.column_dimensions[ws_year.cell(1, col_idx).column_letter].width = 18

    return wb


@extend_schema(
    summary="Выгрузка статистики в Excel",
    description=(
        "Возвращает .xlsx-файл со статистикой посещаемости всех сотрудников: "
        "ФИО, дни и часы за текущую неделю, месяц и год. Доступно только админам."
    ),
    request=None,
    responses={
        200: OpenApiResponse(description="Excel-файл (.xlsx)"),
        403: OpenApiResponse(description="Только для админов"),
    },
    tags=["Reports"],
)
class ExportAttendanceExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        if not request.user.is_admin:
            return Response(
                {"detail": "Только администраторы могут выгружать отчёты"},
                status=status.HTTP_403_FORBIDDEN,
            )

        wb = _build_workbook()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        today = timezone.now().date()
        filename = f"attendance_{today.isoformat()}.xlsx"

        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
